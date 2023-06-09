## Tunnel Vibes
import os

current_dir = os.getcwd()
self = __file__
file_dir = os.path.dirname(self)

os.chdir(file_dir)

import openpyxl
from aux.program.tunnels_cls import Node, Edge
import gc


# Nodes (Buildings)
MC = Node("MC")
SLC = Node("SLC")
PAC = Node("PAC")
STC = Node("STC")
QNC = Node("QNC")
AL = Node("AL")
ML = Node("ML")
PHYS = Node("PHYS")
DP = Node("DP")
DC = Node("DC")
E2 = Node("E2")
E3 = Node("E3")
E5 = Node("E5")
E6 = Node("E6")
E7 = Node("E7")
RCH = Node("RCH")
C1 = Node("C1")
C2 = Node("C2")
DWE = Node("DWE")
HH = Node("HH")
B1 = Node("B1")
B2 = Node("B2")
CIF = Node("CIF")
CPH = Node("CPH")
ECH = Node("ECH")
ESC = Node("ESC")
EV1 = Node("EV1")
EV2 = Node("EV2")
EV3 = Node("EV3")
FH = Node("FH")
GH = Node("GH")
HS = Node("HS")
M3 = Node("M3")
NH = Node("NH")
PAS = Node("PAS")
SCH = Node("SCH")
TC = Node("TC")


objects = gc.get_objects()
Nodes = [node for node in objects if isinstance(node, Node)]



# Edges (connections (tunnels. stairs, etc.))
# edge1 = Edge(MC,PAC)
# edge2 = Edge(MC,SLC)
# edge3 = Edge(DC, MC)
# edge4 = Edge(E3, E5)
# edge5 = Edge(RCH, DWE)
# edge6 = Edge(E2, DWE)
# edge7 = Edge(E3, DC)
# edge8 = Edge(C2, DC)


def make_edges(edges_file):
    # Open the excel file
    workbook = openpyxl.load_workbook(edges_file)
    sheet = workbook.active

    # Create an empty list to store the Person objects
    edges = []

    # Iterate through the rows of data
    for row in range(2, sheet.max_row + 1):
        node1 = sheet.cell(row, 1).value
        node2 = sheet.cell(row, 2).value

        # Create a new Person object and add it to the list
        edge = Edge(next(filter(lambda node: node.name == node1, Nodes)), next(filter(lambda node: node.name == node2, Nodes)))
        edges.append(edge)

    return edges


make_edges("edges.xlsx")

os.chdir(current_dir)


nodes = []
for object in objects:
    if isinstance(object, Node):
        nodes.append(object)

# is_connected(MC,SLC)



